# Для работы программы необходимо сформировать в 1С отчеты и сохранить
# их в Excel с соответствующими названиями: 'К_чеки.xls', 'К_часы.xls', 'К_табель.xls'.
# Целевой файл "Нагрузка кассы.xlsx".
# Программа может давать сбой, если в файле содержатся недопустимые значения
# времени пробития чека, например 21:33. 
import os
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
from openpyxl.writer.excel import save_workbook
# from openpyxl.utils import FORMULAE
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule # Comparison bars
import pandas as pd
import easygui
import statistics
import datetime
import calendar

# Import and process data from files 'Checks'
df = pd.read_excel('К_чеки.xls', header = None)
df.to_excel('К_чеки.xlsx', index = False, header = False)
wbb1 = openpyxl.load_workbook('К_чеки.xlsx')
sh_IP_chex = wbb1.worksheets[0]
chexTotal = sh_IP_chex.max_row-3

# Identifying period (month, year)
dateCell=str(sh_IP_chex['C4'].value)
monthNum = dateCell[3:5]
year = dateCell[6:8]
month_list={'01':'Январь', '02':'Февраль', '03':'Март', '04':'Апрель', '05':'Май', '06':'Июнь',
            '07':'Июль', '08':'Август', '09':'Сентябрь', '10':'Октябрь', '11':'Ноябрь', '12':'Декабрь'}
monthName=month_list[monthNum]
period=str(monthName+' 20'+year)

cl = calendar.Calendar()
dateList = []  # Date and Month for Charts

for i in cl.itermonthdays(int('20'+year), int(monthNum)):
    if i != 0:
        dateList.append(str(i) + " " + monthName[:3])

week_days = {'0': ' Пн', '1':' Вт', '2':' Ср', '3':' Чт', '4':' Пт', '5':' Сб', '6':' Вс'}
d = calendar.monthrange(int('20'+year), int(monthNum))[0] # Adding the week day
week_days2 = {'Пн': 0, 'Вт': 0, 'Ср': 0, 'Чт': 0, 'Пт': 0, 'Сб': 0, 'Вс': 0}

dateList2 = []
dateChexVIA = {} # Dicts for counting checks by days
dateChexRIE = {}
dateChexSHII = {}
week_daysVIA = week_days2.copy()
week_daysRIE = week_days2.copy()
week_daysSHII = week_days2.copy()

for i in dateList:
    dates = i + week_days[str(d)]
    dateChexVIA[dates] = 0
    dateChexRIE[dates] = 0
    dateChexSHII[dates] = 0
    dateList2.append(dates)
    if d == 6:
        d = 0
    else:
        d += 1

# Number and list of girls
entreList = []
girlsVIA = []
girlsRIE = []
girlsSHII = []

for j in range(4, sh_IP_chex.max_row+1):
    ent = sh_IP_chex.cell(row=j, column=5).value
    if ent not in entreList:
        entreList.append(ent)
entreList.sort()

for i in range(4, sh_IP_chex.max_row+1):
    ent = sh_IP_chex.cell(row=i, column=5).value
    if ent == entreList[0]:
        l = girlsVIA
    elif ent == entreList[1]:
        l = girlsRIE
    else:
        l = girlsSHII
    g = sh_IP_chex.cell(row=i, column=6).value
    if g != None and g not in l:
        l.append(g)
    elif g is None:
        x = 'Аноним'
        if x not in l:
            l.append(x)

# wrongNames = ['Вербовская Ирина Александровна', 'Волков Михаил Владимирович', 'Админ', 'Ширяев Илья Игоревич', 'Рейн Инга Эдуардовна']

girlsList = girlsVIA + girlsRIE + girlsSHII  # Complete list of girls


# girlsNumVIA = len(girlsVIA)
# girlsNumRIE = len(girlsRIE)
# girlsNumSHII = len(girlsSHII)


# Number of orders
orderNumVIA = 0
orderNumRIE = 0
orderNumSHII = 0
for i in range(4, sh_IP_chex.max_row+1):
    delChex = sh_IP_chex.cell(row = i, column = 1).value  # Ignoring deleted checks marked with 'X'
    if delChex != 'X':
        order = sh_IP_chex.cell(row = i, column = 7).value
        if order == 1:
            ent = sh_IP_chex.cell(row = i, column = 5).value
            if ent == entreList[0]:
                orderNumVIA += 1
            elif ent == entreList[1]:
                orderNumRIE += 1
            else:
                orderNumSHII += 1
    else:
        continue

# Time Intervals Stats
timeIntervalsTuple = ('08:30','09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30',
                      '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30',
                      '19:00', '19:30', '20:00', '20:30')

# A list of time intervals as keys
timeIntervalsVIA = dict.fromkeys(timeIntervalsTuple, 0)
timeIntervalsRIE = dict.fromkeys(timeIntervalsTuple, 0)
timeIntervalsSHII = dict.fromkeys(timeIntervalsTuple, 0)

for i in range(4, sh_IP_chex.max_row+1):
    ent = sh_IP_chex.cell(row=i, column=5).value
    if ent == entreList[0]:
        li = timeIntervalsVIA
    elif ent == entreList[1]:
        li = timeIntervalsRIE
    else:
        li = timeIntervalsSHII
    t = str(sh_IP_chex.cell(row=i, column=4).value)
    t1 = t[:4]
    t2 = t[:3] + '3'
    if t1 < t2:
        li[t[:3]+'00'] += 1
    else:
        li[t[:3]+'30'] += 1

# Girls' personal statistics
girlsVIAchex = dict.fromkeys(girlsVIA, 0)
girlsVIAorders = dict.fromkeys(girlsVIA, 0)
girlsVIAsums = dict.fromkeys(girlsVIA, 0)
girlsVIAhrs = dict.fromkeys(girlsVIA, 0)
girlsRIEchex = dict.fromkeys(girlsRIE, 0)
girlsRIEorders = dict.fromkeys(girlsRIE, 0)
girlsRIEsums = dict.fromkeys(girlsRIE, 0)
girlsRIEhrs = dict.fromkeys(girlsRIE, 0)
girlsSHIIchex = dict.fromkeys(girlsSHII, 0)
girlsSHIIorders = dict.fromkeys(girlsSHII, 0)
girlsSHIIsums = dict.fromkeys(girlsSHII, 0)
girlsSHIIhrs = dict.fromkeys(girlsSHII, 0)
# girlsHrs = dict.fromkeys(girlsList, 0)

chexNumVIA = 0
chexNumRIE = 0
chexNumSHII = 0
checkSumVIA = []
checkSumRIE = []
checkSumSHII = []
maxValVIA = 0
maxValRIE = 0
maxValSHII = 0

for i in range(4, sh_IP_chex.max_row+1):
    delChex = sh_IP_chex.cell(row=i, column=1).value
    if delChex != 'X':
        name = sh_IP_chex.cell(row=i, column=6).value
        if name is None:
            name = 'Аноним'
        order = sh_IP_chex.cell(row=i, column=7).value
        checkSum = sh_IP_chex.cell(row=i, column=8).value
        ent = sh_IP_chex.cell(row=i, column=5).value

        if ent == entreList[0]:
            girlsVIAchex[name] += 1
            girlsVIAorders[name] += order
            girlsVIAsums[name] += checkSum
            chexNumVIA +=1
            checkSumVIA.append(checkSum)
            if checkSum > maxValVIA:  # The maximum check sum
                maxValVIA = int(checkSum)
                maxValVIADate = sh_IP_chex.cell(row=i, column=3).value
            for cd in dateChexVIA.keys():  # Counts checks by entr. and days
                checkDate = int(sh_IP_chex.cell(row=i, column=3).value[:2])
                if checkDate == int(cd[:2]):
                    dateChexVIA[cd] += 1
        elif ent == entreList[1]:
            girlsRIEchex[name] += 1
            girlsRIEorders[name] += order
            girlsRIEsums[name] += checkSum
            chexNumRIE +=1
            checkSumRIE.append(checkSum)
            if checkSum > maxValRIE:  # The maximum check sum
                maxValRIE = int(checkSum)
                maxValRIEDate = sh_IP_chex.cell(row=i, column=3).value
            for cd in dateChexRIE.keys():  # Counts checks by entr. and days
                checkDate = int(sh_IP_chex.cell(row=i, column=3).value[:2])
                if checkDate == int(cd[:2]):
                    dateChexRIE[cd] += 1
        elif ent == entreList[2]:
            girlsSHIIchex[name] += 1
            girlsSHIIorders[name] += order
            girlsSHIIsums[name] += checkSum  # No int()
            chexNumSHII +=1
            checkSumSHII.append(checkSum)
            if checkSum > maxValSHII:  # The maximum check sum
                maxValSHII = int(checkSum)
                maxValSHIIDate = sh_IP_chex.cell(row=i, column=3).value
            for cd in dateChexSHII.keys():  # Counts checks by entr. and days
                checkDate = int(sh_IP_chex.cell(row=i, column=3).value[:2])
                if checkDate == int(cd[:2]):
                    dateChexSHII[cd] += 1
    else:
        continue

for ddt in dateChexVIA.keys():  # Weekdays check number VIA
    wd = str(ddt[-2:])
    week_daysVIA[wd] += dateChexVIA[ddt]
    week_days2[wd] += dateChexVIA[ddt]
for ddt in dateChexRIE.keys():  # Weekdays check number RIE
    wd = str(ddt[-2:])
    week_daysRIE[wd] += dateChexRIE[ddt]
    week_days2[wd] += dateChexRIE[ddt]
for ddt in dateChexSHII.keys():  # Weekdays check number SHII
    wd = str(ddt[-2:])
    week_daysSHII[wd] += dateChexSHII[ddt]
    week_days2[wd] += dateChexSHII[ddt]  # Totals by weekdays

# Mean, median, mode
def mean(sumList):
    return sum(sumList) / len(sumList)
checkMeanVIA = int(mean(checkSumVIA))
checkMeanRIE = int(mean(checkSumRIE))
checkMeanSHII = int(mean(checkSumSHII))

checkMedianVIA = statistics.median(checkSumVIA)
checkMedianRIE = statistics.median(checkSumRIE)
checkMedianSHII = statistics.median(checkSumSHII)
checkQuartileVIA = statistics.quantiles(checkSumVIA, n=4)
checkQuartileRIE = statistics.quantiles(checkSumRIE, n=4)
checkQuartileSHII = statistics.quantiles(checkSumSHII, n=4)


checkSumVIArounded = []  # Rounding check sums for finding the mode value
for i in checkSumVIA:
    v = round(i/25)*25
    checkSumVIArounded.append(v)

checkSumRIErounded = []
for i in checkSumRIE:
    v = round(i/25)*25
    checkSumRIErounded.append(v)

checkSumSHIIrounded = []
for i in checkSumSHII:
    v = round(i/25)*25
    checkSumSHIIrounded.append(v)

checkModeVIA = statistics.mode(checkSumVIArounded)
checkModeRIE = statistics.mode(checkSumRIErounded)
checkModeSHII = statistics.mode(checkSumSHIIrounded)

numModeVIA = 0  # Occurrences
for m in checkSumVIArounded:
    if m == checkModeVIA:
        numModeVIA += 1
perModeVIA = round(numModeVIA * 100 / len(checkSumVIArounded), 1)
occurVIA = str(numModeVIA) + ', ' + str(perModeVIA) + '%'

numModeRIE = 0
for m in checkSumRIErounded:
    if m == checkModeRIE:
        numModeRIE += 1
perModeRIE = round(numModeRIE * 100 / len(checkSumVIArounded), 1)
occurRIE = str(numModeRIE) + ', ' + str(perModeRIE) + '%'

numModeSHII = 0
for m in checkSumSHIIrounded:
    if m == checkModeSHII:
        numModeSHII += 1
perModeSHII = round(numModeSHII * 100 / len(checkSumSHIIrounded), 1)
occurSHII = str(numModeSHII) + ', ' + str(perModeSHII) + '%'

# Dicts of moda intervals
# def modaIntVIA(n, v):    # Funcion sorts out checkSumVIArounded into Moda Intervals
#     if v <= modaIntervals[n]:
#         if n-1 < 1:
#             return
#         else:
#             if v > modaIntervals[n-1]:
#                 modaDictVIA[n-1] += 1
#             else:
#                 modaIntVIA(n-1, v)
#     else:
#         if v > modaIntervals[n+1]:
#             if n+2 > 12:
#                 modaDictVIA[n+1] += 1
#             else:
#                 modaIntVIA(n+1, v)


def modaInt(v, dict):
    if v <= 50:
        dict[50] += 1
    elif v > 50 and v <= 75:
        dict[75] += 1
    elif v > 75 and v <= 100:
        dict[100] += 1
    elif v > 100 and v <= 250:
        dict[250] += 1
    elif v > 250 and v <= 500:
        dict[500] += 1
    elif v > 500 and v <= 1000:
        dict[1000] += 1
    elif v > 1000 and v <= 1500:
        dict[1500] += 1
    elif v > 1500 and v <= 2000:
        dict[2000] += 1
    elif v > 2000 and v <= 3000:
        dict[3000] += 1
    elif v > 3000 and v <= 5000:
        dict[5000] += 1
    elif v > 5000 and v <= 10000:
        dict[10000] += 1
    elif v > 10000 and v <= 20000:
        dict[20000] += 1
    elif v > 20000 and v <= 50000:
        dict[50000] += 1
    elif v > 50000 and v <= 75000:
        dict[75000] += 1
    elif v > 75000 and v <= 100000:
        dict[100000] += 1
    else:
        dict[100001] += 1

modaIntervals = [50, 75, 100, 250, 500, 1000, 1500, 2000, 3000, 5000, 10000, 20000, 50000, 75000, 100000, 100001]
modaDictVIA = {k: 0 for k in modaIntervals}
modaDictRIE = {k: 0 for k in modaIntervals}
modaDictSHII = {k: 0 for k in modaIntervals}

for v in checkSumVIA:
    modaInt(v, modaDictVIA)
for v in checkSumRIE:
    modaInt(v, modaDictRIE)
for v in checkSumSHII:
    modaInt(v, modaDictSHII)



# Sorting girls...chex by values
# valSort_girlsVIAchex = sorted(girlsVIAchex.values(), reverse=True)
sortedgirlsVIAchexTuples = sorted(girlsVIAchex.items(), key=lambda item: item[1], reverse=True)
girlsVIAchexSorted = {k:v for k,v in sortedgirlsVIAchexTuples}
# girlsVIAchexSorted = {}
# for i in valSort_girlsVIAchex:
#     for k in girlsVIAchex.keys():
#         if girlsVIAchex[k] == i:
#             girlsVIAchexSorted[k] = girlsVIAchex[k]

sortedgirlsRIEchexTuples = sorted(girlsRIEchex.items(), key=lambda item: item[1], reverse=True)
girlsRIEchexSorted = {k:v for k,v in sortedgirlsRIEchexTuples}

sortedgirlsSHIIchexTuples = sorted(girlsSHIIchex.items(), key=lambda item: item[1], reverse=True)
girlsSHIIchexSorted = {k:v for k,v in sortedgirlsSHIIchexTuples}

wbb1.close()
os.remove('К_чеки.xlsx')
# ++++++++++++++++++++++++++++============================================================

# Import and process data from file 'Hours'

# df = pd.read_excel('К_часы.xls', header = None)
# df.to_excel('К_часы.xlsx', index = False, header = False)
# wbb3 = openpyxl.load_workbook('К_часы.xlsx')
# hours_IP = wbb3.worksheets[0]
# hours_IP.max_row-3
# cas = 'Кассиры'
# hours_VIA = 0
# hours_RIE = 0
# hours_SHII = 0
# for i in range (1, hours_IP.max_row):
#     dptName = str(hours_IP.cell(row=i, column=5).value)
#     if cas in dptName and str(entreList[0]) in dptName:
#         hours_VIA += hours_IP.cell(row=i, column=6).value
#     elif cas in dptName and str(entreList[1]) in dptName:
#         hours_RIE += hours_IP.cell(row=i, column=6).value
#     elif cas in dptName and str(entreList[2]) in dptName:
#         hours_SHII += hours_IP.cell(row=i, column=6).value
# wbb3.close()
# os.remove('К_часы.xlsx')
# ++++++++++++++++++++++++++++============================================================

# Import and process data from file 'Time table'
girlsHrs = {}
df = pd.read_excel('К_табель.xls', header = None)
df.to_excel('К_табель.xlsx', index = False, header = False)
wbb5 = openpyxl.load_workbook('К_табель.xlsx')
timeTable = wbb5.worksheets[0]
hrsCol = timeTable.max_column
for i in range(9, timeTable.max_row+1):
    name = timeTable.cell(row=i, column=1).value
    hrs = float(timeTable.cell(row=i, column=hrsCol).value)
    girlsHrs[name] = hrs


# hours_VIA
# hours_RIE
# hours_SHII


wbb5.close()
os.remove('К_табель.xlsx')

# ++++++++++++++++++++++++++++============================================================
# Import and process data from file 'K_bonus'

# girlsBon = dict.fromkeys(girlsHrs, 0)
girlsBon = {x: [0, 0] for x in girlsHrs.keys()}  # need to generate a dict with a list of two entries


df = pd.read_excel('К_бонус.xls', header = None)
df.to_excel('К_бонус.xlsx', index = False, header = False)
wbb6 = openpyxl.load_workbook('К_бонус.xlsx')
shBon = wbb6.worksheets[0]
bonCol = shBon.max_column

for i in range(4, shBon.max_row+1):
    if shBon.cell(row=i, column=1).value != 'X':
        name = shBon.cell(row=i, column=6).value
        if name not in girlsHrs.keys():
            continue
        else:
            bon = shBon.cell(row=i, column=bonCol).value
            if shBon.cell(row=i, column=7).value  != 'Чек на возврат':
                girlsBon[name][0] += 1
                girlsBon[name][1] += round(bon, 0)
            else:
                girlsBon[name][0] -= 1
                girlsBon[name][1] -= round(bon, 0)
    else:
        continue

wbb6.close()
os.remove('К_бонус.xlsx')


# ----------------------------------------------------------------------------------------
# Cleansing the lists and dicts
for i in girlsHrs.keys():
    if girlsHrs[i] == 0:
        girlsHrs.popitem()

girlsVIA2 = girlsVIA.copy()  # These lists are free of wrong names like entrepreneurs' names, Admin, etc.
girlsRIE2 = girlsRIE.copy()
girlsSHII2 = girlsSHII.copy()

for i in girlsVIA:
    if i not in girlsHrs:
        girlsVIA2.remove(i)

for i in girlsRIE:
    if i not in girlsHrs:
        girlsRIE2.remove(i)

for i in girlsSHII:
    if i not in girlsHrs:
        girlsSHII2.remove(i)

girlsNumVIA = len(girlsVIA2)
girlsNumRIE = len(girlsRIE2)
girlsNumSHII = len(girlsSHII2)





# ++++++++++++++++++++++++++++============================================================

workloadFile = 'Статистика кассы.xlsx'
try:
    wbb4 = openpyxl.load_workbook(workloadFile)
except:
    openpyxl.Workbook()
wsh4 = wbb4.create_sheet(period)
wbb4.active = wsh4

# Cell style
boldFont = Font(bold=True, size=7)
plainFont = Font(bold=False, size=9)
alignCenter = Alignment(horizontal='center', vertical='center')
alignLeft = Alignment(horizontal='left', vertical='center')
borderLines = Side(border_style='thin', color='000000')
squareBorder = Border(top=borderLines,
                      bottom=borderLines,
                      right=borderLines,
                      left=borderLines)
cellFillYellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
cellFillGreenish = PatternFill(start_color='CCFF99', end_color='CCFF99', fill_type='solid')

# Drawing the tables
# Table 1
for i in range(3, 7):
    for j in range(1, 18):
        wsh4.cell(row=i, column=j).border = squareBorder
        wsh4.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        # wsh4.cell(row=i, column=j).alignment = alignCenter
        if i == 3:
            wsh4.cell(row=i, column=j).font = boldFont
            wsh4.cell(row=i, column=j).fill = cellFillYellow
        if j == 1 and i != 3:
            wsh4.cell(row=i, column=j).fill = cellFillGreenish


wsh4.column_dimensions['A'].width = 23
cellList = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'T']
for j in range(3, 7):
    wsh4.row_dimensions[j].height = 21
for i in cellList:
    wsh4.column_dimensions[i].width = 11
# wsh4.column_dimensions['M'].width = 12
# wsh4.column_dimensions['N'].width = 12
wsh4['E2'].font = boldFont
wsh4['A3'] = period


# Table 2
for i in range(9, 13):
    for j in range(1, 18):
        wsh4.cell(row=i, column=j).border = squareBorder
        wsh4.cell(row=i, column=j).alignment = alignCenter
        if i == 9:
            wsh4.cell(row=i, column=j).font = boldFont
            wsh4.cell(row=i, column=j).fill = cellFillYellow
        if j == 1 and i != 9:
            wsh4.cell(row=i, column=j).fill = cellFillGreenish
for j in range(9, 13):
    wsh4.row_dimensions[j].height = 14


# Table 3  (border drawing part is below, where data is inserted
wsh4.row_dimensions[15].height = 12
for j in range (3, 10):
    wsh4.cell(row=15, column=j).border = squareBorder
    wsh4.cell(row=15, column=j).fill = cellFillYellow
    wsh4.cell(row=15, column=j).alignment = alignCenter


# Filling the tables
# wsh4['A4'] = '  ИП Вербовская И.А.'
# wsh4['A5'] = '  ИП Рейн И.Э.'
# wsh4['A6'] = '  ИП Ширяев И.И.'

ip_names = ['  ИП Вербовская И.А.', '  ИП Рейн И.Э.', '  ИП Ширяев И.И.']
rowIP = 4
for ip in ip_names:
    wsh4.cell(row=rowIP, column=1).value = ip
    wsh4.cell(row=rowIP, column=1).alignment = alignLeft
    rowIP += 1
rowIP += 3
for ip in ip_names:
    wsh4.cell(row=rowIP, column=1).value = ip
    wsh4.cell(row=rowIP, column=1).alignment = alignLeft # wsh4['A4'].alignment = alignLeft
    rowIP += 1

headerList = ['Кассиров', 'Чеков', 'Заказов', 'Всего часов', 'Чек / час',
              'Чек / сотр.', 'Заказ / сотр.', 'Час / сотр.', 'Среднее', 'Нижн. квартиль', 'Медиана',
              'Верхн. квартиль', 'Мода', 'Частота', 'Макс. чек', 'Дата чека']

headerList2 = ['Модальные интервалы', '<= 50', '<= 75', '<= 100', '<= 250', '<= 500', '<=1 000', '<=1 500',
               '<=2 000', '<= 3 000', '<= 5 000',
               '<= 10 000', '<= 20 000', '< = 50 000', '<= 75 000', '<= 100 000', '> 100 000']

headerList3 = ['Кол-во чек.', 'Заказов', 'Часов', 'Чек/час', 'СБП Кол-во', 'СБП Сумма', 'Бонус']

# Table headers
rowT = 3
col = 2
for i in headerList:
    wsh4.cell(row=rowT, column=col).value = i
    col += 1

for i in range(4, 7):
    for j in range(1, 18):
        wsh4.cell(row=i, column=j).font = plainFont

rowT += 6
col = 1
for i in headerList2:
    wsh4.cell(row=rowT, column=col).value = i
    col += 1

rowT += 6
col = 3
for i in headerList3:
    wsh4.cell(row=rowT, column=col).value = i
    col += 1

# Inserting data
wsh4['B4'] = girlsNumVIA
wsh4['B5'] = girlsNumRIE
wsh4['B6'] = girlsNumSHII
wsh4['C4'] = chexNumVIA
wsh4['C5'] = chexNumRIE
wsh4['C6'] = chexNumSHII
wsh4['D4'] = orderNumVIA
wsh4['D5'] = orderNumRIE
wsh4['D6'] = orderNumSHII
# wsh4['E4'] = hours_VIA
# wsh4['E5'] = hours_RIE
# wsh4['E6'] = hours_SHII
# wsh4['F4'] = round(chexNumVIA / hours_VIA, 1) #'=C4 / E4'
# wsh4['F5'] = round(chexNumRIE / hours_RIE, 1) #'=C5 / E5'
# wsh4['F6'] = round(chexNumSHII / hours_SHII, 1) #'=C6 / E6'
wsh4['G4'] = round(chexNumVIA /  girlsNumVIA, 0)
wsh4['G5'] = round(chexNumRIE /  girlsNumRIE, 0)
wsh4['G6'] = round(chexNumSHII /  girlsNumSHII, 0)
wsh4['H4'] = round(orderNumVIA /  girlsNumVIA, 1)
wsh4['H5'] = round(orderNumRIE /  girlsNumRIE, 1)
wsh4['H6'] = round(orderNumSHII /  girlsNumSHII, 1)
# wsh4['I4'] = round(hours_VIA / girlsNumVIA, 1) #'=ОКРУГЛ(E4/B4;1)'
# wsh4['I5'] = round(hours_RIE / girlsNumRIE, 1) # '=ОКРУГЛ(E5/B5;1)'
# wsh4['I6'] = round(hours_SHII / girlsNumSHII, 1) #'=ОКРУГЛ(E6/B6;1)'
wsh4['J4'] = checkMeanVIA
wsh4['J5'] = checkMeanRIE
wsh4['J6'] = checkMeanSHII

wsh4['K4'] = round(checkQuartileVIA[0], 0)
wsh4['K5'] = round(checkQuartileRIE[0], 0)
wsh4['K6'] = round(checkQuartileSHII[0], 0)

wsh4['L4'] = round(checkMedianVIA, 0)
wsh4['L5'] = round(checkMedianRIE, 0)
wsh4['L6'] = round(checkMedianSHII, 0)

wsh4['M4'] = round(checkQuartileVIA[2], 0)
wsh4['M5'] = round(checkQuartileRIE[2], 0)
wsh4['M6'] = round(checkQuartileSHII[2], 0)

wsh4['N4'] = checkModeVIA
wsh4['N5'] = checkModeRIE
wsh4['N6'] = checkModeSHII
wsh4['O4'] = occurVIA
wsh4['O5'] = occurRIE
wsh4['O6'] = occurSHII


wsh4['P4'] = maxValVIA
wsh4['P5'] = maxValRIE
wsh4['P6'] = maxValSHII
wsh4['Q4'] = maxValVIADate
wsh4['Q5'] = maxValRIEDate
wsh4['Q6'] = maxValSHIIDate

# wsh4.cell(row=4, column=3).value="{:,}".format(chexNumVIA).replace(',', ' ')
# wsh4.cell(row=5, column=3).value="{:,}".format(chexNumRIE).replace(',', ' ')
# wsh4.cell(row=6, column=3).value="{:,}".format(chexNumSHII).replace(',', ' ')
#
# wsh4.cell(row=4, column=13).value = "{:,}".format(maxValVIA).replace(',', ' ')
# wsh4.cell(row=5, column=13).value = "{:,}".format(maxValRIE).replace(',', ' ')
# wsh4.cell(row=6, column=13).value = "{:,}".format(maxValSHII).replace(',', ' ')


# The easiest way to format Excel cells, just cool!
for r in range(16, 52):
    wsh4[f'C{r}'].number_format = '# ##0'
    wsh4[f'E{r}'].number_format ='0.0'
    wsh4[f'H{r}'].number_format ='# ##0'

for q in range(2, 18):
    for p in range(10, 13):
        wsh4.cell(row=p, column=q).number_format = '# ##0'

for r in range (4, 7):
    wsh4[f'C{r}'].number_format = '# ##0'
    wsh4[f'E{r}'].number_format = '# ##0.0'
    wsh4[f'F{r}'].number_format = '0.0'
    wsh4[f'G{r}'].number_format = '# ##0'
    wsh4[f'J{r}'].number_format = '# ##0'
    wsh4[f'O{r}'].number_format = '# ##0'
    wsh4[f'P{r}'].number_format = '# ##0'

# Moda intervals
col = 2
for i in modaIntervals:
    wsh4.cell(row=10, column=col).value = modaDictVIA[i]
    wsh4.cell(row=11, column=col).value = modaDictRIE[i]
    wsh4.cell(row=12, column=col).value = modaDictSHII[i]
    col += 1




# Girls' personal stats part

data_bar_rule = DataBarRule(start_type="num", start_value=1, end_type="num", end_value="15", color="0000FF00")  # Green
wsh4.conditional_formatting.add("F16:F48", data_bar_rule)

# The 3 parts of Table 3. The last one is different from the first two.

# Function for returning dict. key by value
def getKeyVIA(val):
    for key, value in girlsVIAchexSorted.items():
        if val == value:
            return key

hours_VIA = 0
row = rowT + 1
for i in girlsVIAchexSorted.values():  # Inserting list sorted by check qty
    wsh4.cell(row=row, column=3).value = int(i) # Inserting check qty
    wsh4.cell(row=row, column=1).value = getKeyVIA(i) # Inserting names
    wsh4.cell(row=row, column=4).value = girlsVIAorders.get(getKeyVIA(i))  # Inserting order qty
    wsh4.cell(row=row, column=5).value = girlsHrs.get(getKeyVIA(i)) # Inserting person's total time
    if getKeyVIA(i) in girlsHrs:
        wsh4.cell(row=row, column=6).value = round(i / girlsHrs.get(getKeyVIA(i)), 1)
        hours_VIA += girlsHrs.get(getKeyVIA(i))
    else:
        continue
    wsh4.cell(row=row, column=7).value = girlsBon.get(getKeyVIA(i))[0]  # Inserting bonus
    wsh4.cell(row=row, column=8).value = girlsBon.get(getKeyVIA(i))[1]
    wsh4.cell(row=row, column=9).value = round(girlsBon.get(getKeyVIA(i))[1] * 0.1 / 100, 0)
    row += 1


# Function for returning dict. key by value
def getKeyRIE(val):
    for key, value in girlsRIEchexSorted.items():
        if val == value:
            return key

hours_RIE = 0
row +=2
for i in girlsRIEchexSorted.values():  # Inserting list sorted by check qty
    wsh4.cell(row=row, column=3).value = i
    wsh4.cell(row=row, column=1).value = getKeyRIE(i)
    wsh4.cell(row=row, column=4).value = girlsRIEorders.get(getKeyRIE(i))
    wsh4.cell(row=row, column=5).value = girlsHrs.get(getKeyRIE(i))  # Inserting person's total time
    if getKeyRIE(i) in girlsHrs:
        wsh4.cell(row=row, column=6).value = round(i / girlsHrs.get(getKeyRIE(i)), 1)
        hours_RIE += girlsHrs.get(getKeyRIE(i))
    else:
        continue
    wsh4.cell(row=row, column=7).value = girlsBon.get(getKeyRIE(i))[0]  # Inserting bonus
    wsh4.cell(row=row, column=8).value = girlsBon.get(getKeyRIE(i))[1]
    wsh4.cell(row=row, column=9).value = round(girlsBon.get(getKeyRIE(i))[1] * 0.1 / 100, 0)
    row += 1

# Function for returning dict. key by value
# def getKeySHII(val):
#     for key, value in girlsSHIIchexSorted.items():
#         if val == value:
#             return key

hours_SHII = 0
row +=2
for i in girlsSHIIchexSorted.keys():  # Inserting list sorted by check qty
    wsh4.cell(row=row, column=1).value = i
    wsh4.cell(row=row, column=3).value = girlsSHIIchexSorted[i]
    # wsh4.cell(row=row, column=1).value = getKeySHII(i)  # Delete?
    wsh4.cell(row=row, column=4).value = girlsSHIIorders[i]
    if i in girlsHrs:
        wsh4.cell(row=row, column=5).value = round(girlsHrs[i], 1)  # Inserting person's total time
        wsh4.cell(row=row, column=6).value = round(girlsSHIIchexSorted[i] / girlsHrs[i], 1)
        hours_SHII += girlsHrs[i]
    else:
        continue
    wsh4.cell(row=row, column=7).value = girlsBon[i][0]  # Inserting bonus
    wsh4.cell(row=row, column=8).value = girlsBon[i][1]
    wsh4.cell(row=row, column=9).value = round(girlsBon[i][1] * 0.1 / 100, 0)
    row += 1

# Adding hours data
wsh4['E4'] = hours_VIA
wsh4['E5'] = hours_RIE
wsh4['E6'] = hours_SHII
wsh4['F4'] = round(chexNumVIA / hours_VIA, 1) #'=C4 / E4'
wsh4['F5'] = round(chexNumRIE / hours_RIE, 1) #'=C5 / E5'
wsh4['F6'] = round(chexNumSHII / hours_SHII, 1) #'=C6 / E6'
wsh4['I4'] = round(hours_VIA / girlsNumVIA, 1) #'=ОКРУГЛ(E4/B4;1)'
wsh4['I5'] = round(hours_RIE / girlsNumRIE, 1) # '=ОКРУГЛ(E5/B5;1)'
wsh4['I6'] = round(hours_SHII / girlsNumSHII, 1) #'=ОКРУГЛ(E6/B6;1)'


# Drawing borders for the table 3
for i in range(16, 55):
    if wsh4.cell(row=i, column=1).value != None:
        for j in range(1, 10):
            wsh4.cell(row=i, column=j).border = squareBorder



# Time Intervals Chart Data
row += 20
row2 = row
wsh4.cell(row=row, column=19).value = 'БД1'
wsh4.cell(row=row, column=20).value = 'БД3'
wsh4.cell(row=row, column=21).value = 'БД4'
row += 1

for i in timeIntervalsVIA.keys():
    j = timeIntervalsVIA[i]
    k = timeIntervalsRIE[i]
    l = timeIntervalsSHII[i]
    wsh4.cell(row=row, column=18).value = i
    wsh4.cell(row=row, column=19).value = j
    wsh4.cell(row=row, column=20).value = k
    wsh4.cell(row=row, column=21).value = l
    row += 1
row -= 1

# Time Intervals Bar Chart
timeIntervalsChart = BarChart()
timeIntervalsChart.title = 'Статистика по часам'
chartCats = Reference(worksheet=wsh4,
                      min_row=row2+1, max_row=row,
                      min_col=18, max_col=18)
chartData = Reference(worksheet=wsh4,
                      min_row=row2, max_row=row,
                      min_col=19, max_col=21)
timeIntervalsChart.add_data(chartData, titles_from_data=True)
timeIntervalsChart.x_axis.title = "Временные интервалы"
timeIntervalsChart.y_axis.title = "Кол-во чеков по ИП"
timeIntervalsChart.width = 29
timeIntervalsChart.height = 8
wsh4.add_chart(timeIntervalsChart, "A76")
timeIntervalsChart.set_categories(chartCats)

# Stats by Dates Chart Data
row += 3
row2 = row
wsh4.cell(row=row, column=19).value = 'ВИА'
wsh4.cell(row=row, column=20).value = 'РИЭ'
wsh4.cell(row=row, column=21).value = 'ШИИ'
row += 1

for i in dateChexVIA.keys():
    j = dateChexVIA[i]
    k = dateChexRIE[i]
    l = dateChexSHII[i]
    wsh4.cell(row=row, column=18).value = i
    wsh4.cell(row=row, column=19).value = j
    wsh4.cell(row=row, column=20).value = k
    wsh4.cell(row=row, column=21).value = l
    row += 1
row -= 1

# Stats by Dates Chart
datesChart = BarChart()
datesChart.title = 'Статистика за месяц по дням'
chartCats = Reference(worksheet=wsh4,
                      min_row=row2+1, max_row=row,
                      min_col=18, max_col=18)
chartData = Reference(worksheet=wsh4,
                      min_row=row2, max_row=row,
                      min_col=19, max_col=21)
datesChart.add_data(chartData, titles_from_data=True)
datesChart.x_axis.title = "Дни месяца, дни недели"
datesChart.y_axis.title = "Кол-во чеков по ИП"
datesChart.width = 29
datesChart.height = 8
wsh4.add_chart(datesChart, "A101")
datesChart.set_categories(chartCats)

# Week Days Chart Data
row = 53
row2 = row
wsh4.cell(row=row, column=19).value = 'ВИА'
wsh4.cell(row=row, column=20).value = 'РИЭ'
wsh4.cell(row=row, column=21).value = 'ШИИ'
row += 1

for i in week_days2.keys():
    j = week_daysVIA[i]
    k = week_daysRIE[i]
    l = week_daysSHII[i]
    m = week_days2[i]
    wsh4.cell(row=row, column=18).value = i
    wsh4.cell(row=row, column=19).value = j
    wsh4.cell(row=row, column=20).value = k
    wsh4.cell(row=row, column=21).value = l
    wsh4.cell(row=row, column=22).value = m
    row += 1
row -= 1

# Week Days Chart
week_daysChart = BarChart()
week_daysChart.title = 'Статистика по дням недели'
chartCats = Reference(worksheet=wsh4,
                      min_row=row2+1, max_row=row,
                      min_col=18, max_col=18)
chartData = Reference(worksheet=wsh4,
                      min_row=row2, max_row=row,
                      min_col=19, max_col=21)
week_daysChart.add_data(chartData, titles_from_data=True)
week_daysChart.x_axis.title = "Дни недели"
week_daysChart.y_axis.title = "Кол-во чеков по ИП"
week_daysChart.width = 14.7
wsh4.add_chart(week_daysChart, "A53")
week_daysChart.set_categories(chartCats)

# All Entrepreneurs Week Days Chart
week_daysChart = BarChart()
week_daysChart.title = 'Статистика по дням недели суммарно'
chartCats = Reference(worksheet=wsh4,
                      min_row=row2+1, max_row=row,
                      min_col=18, max_col=18)
chartData = Reference(worksheet=wsh4,
                      min_row=row2, max_row=row,
                      min_col=22, max_col=22)
week_daysChart.add_data(chartData, titles_from_data=True)
week_daysChart.x_axis.title = "Дни недели"
week_daysChart.y_axis.title = "Кол-во чеков суммарно"
week_daysChart.width = 14.1
wsh4.add_chart(week_daysChart, "I53")
week_daysChart.set_categories(chartCats)

wsh4.print_area = 'A1:Q120'

wbb4.save(str(workloadFile))
wbb4.close()

print('Accomplished')
