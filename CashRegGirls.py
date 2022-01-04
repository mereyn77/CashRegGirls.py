import os

import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
from openpyxl.writer.excel import save_workbook
# from openpyxl.utils import FORMULAE
from openpyxl.chart import BarChart, Reference
import pandas as pd
import easygui
import statistics
import calendar



# Import and process data from files 'Checks'
df = pd.read_excel('ИП_чеки.xls', header = None)
df.to_excel('ИП_чеки.xlsx', index = False, header = False)
wbb1 = openpyxl.load_workbook('ИП_чеки.xlsx')
sh_IP_chex = wbb1.worksheets[0]
chexTotal = sh_IP_chex.max_row-3

# Identifying period (month, year)
dateCell=str(sh_IP_chex['C4'].value)
monthNum = dateCell[3:5]
year = dateCell[6:8]
month_list={'01':'Январь', '02':'Февраль', '03':'Март', '04':'Апрель', '05':'Май', '06':'Июнь',
            '07':'Июль', '08':'Август', '09':'Сентябрь', '10':'Октябрь', '11':'Ноябрь', '12':'Декабрь'}
monthName=month_list[monthNum]
period=str(monthName+' 20'+year)

cl = calendar.Calendar()

dateList = []  # Date and Month for Charts

for i in cl.itermonthdays(int('20'+year), int(monthNum)):
    if i != 0:
        dateList.append(str(i) + " " + monthName[:3])
        # dateListx.append(str(i))

week_days = {'0': ' Пн', '1':' Вт', '2':' Ср', '3':' Чт', '4':' Пт', '5':' Сб', '6':' Вс'}
d = calendar.monthrange(int('20'+year), int(monthNum))[0] # Adding the week day

dateList2 = []
dateChexVIA = {} # Dicts for counting checks by days
dateChexRIE = {}
dateChexSHII = {}

for i in dateList:
    dates = i + week_days[str(d)]
    dateChexVIA[dates] = 0
    dateChexRIE[dates] = 0
    dateChexSHII[dates] = 0
    dateList2.append(dates)
    if d == 6:
        d = 0
    else:
        d += 1

# dateChexTuple = dateList2

# Number and list of girls
entreList = []
girlsVIA = []
girlsRIE = []
girlsSHII = []

for j in range(4, sh_IP_chex.max_row+1):
    ent = sh_IP_chex.cell(row=j, column=5).value
    if ent not in entreList:
        entreList.append(ent)
entreList.sort()

for i in range(4, sh_IP_chex.max_row+1):
    ent = sh_IP_chex.cell(row=i, column=5).value
    if ent == entreList[0]:
        l = girlsVIA
    elif ent == entreList[1]:
        l = girlsRIE
    else:
        l = girlsSHII
    g = sh_IP_chex.cell(row=i, column=6).value
    if g != None and g not in l:
        l.append(g)
    elif g is None:
        x = 'Аноним'
        if x not in l:
            l.append(x)

girlsNumVIA = len(girlsVIA)
girlsNumRIE = len(girlsRIE)
girlsNumSHII = len(girlsSHII)

# Number of orders
orderNumVIA = 0
orderNumRIE = 0
orderNumSHII = 0
for i in range(4, sh_IP_chex.max_row+1):
    order = sh_IP_chex.cell(row = i, column = 7).value
    if order == 1:
        ent = sh_IP_chex.cell(row = i, column = 5).value
        if ent == entreList[0]:
            orderNumVIA += 1
        elif ent == entreList[1]:
            orderNumRIE += 1
        else:
            orderNumSHII += 1

# Time Intervals Stats
timeIntervalsTuple = ('08:30','09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30',
                      '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30',
                      '19:00', '19:30', '20:00')

# A list of time intervals as keys
timeIntervalsVIA = dict.fromkeys(timeIntervalsTuple, 0)
timeIntervalsRIE = dict.fromkeys(timeIntervalsTuple, 0)
timeIntervalsSHII = dict.fromkeys(timeIntervalsTuple, 0)

for i in range(4, sh_IP_chex.max_row+1):
    ent = sh_IP_chex.cell(row=i, column=5).value
    if ent == entreList[0]:
        li = timeIntervalsVIA
    elif ent == entreList[1]:
        li = timeIntervalsRIE
    else:
        li = timeIntervalsSHII
    t = str(sh_IP_chex.cell(row=i, column=4).value)
    t1 = t[:4]
    t2 = t[:3] + '3'
    if t1 < t2:
        li[t[:3]+'00'] += 1
    else:
        li[t[:3]+'30'] += 1


# VIA Girls' personal statistics
girlsVIAchex = dict.fromkeys(girlsVIA, 0)
girlsVIAorders = dict.fromkeys(girlsVIA, 0)
girlsVIAsums = dict.fromkeys(girlsVIA, 0)
girlsRIEchex = dict.fromkeys(girlsRIE, 0)
girlsRIEorders = dict.fromkeys(girlsRIE, 0)
girlsRIEsums = dict.fromkeys(girlsRIE, 0)
girlsSHIIchex = dict.fromkeys(girlsSHII, 0)
girlsSHIIorders = dict.fromkeys(girlsSHII, 0)
girlsSHIIsums = dict.fromkeys(girlsSHII, 0)
chexNumVIA = 0
chexNumRIE = 0
chexNumSHII = 0
checkSumVIA = []
checkSumRIE = []
checkSumSHII = []
maxValVIA = 0
maxValRIE = 0
maxValSHII = 0

for i in range(4, sh_IP_chex.max_row+1):
    name = sh_IP_chex.cell(row=i, column=6).value
    if name is None:
        name = 'Аноним'
    order = sh_IP_chex.cell(row=i, column=7).value
    checkSum = sh_IP_chex.cell(row=i, column=8).value
    ent = sh_IP_chex.cell(row=i, column=5).value
    # dd = str(sh_IP_chex.cell(row=i, column=3).value[:3])
    if ent == entreList[0]:
        girlsVIAchex[name] += 1
        girlsVIAorders[name] += order
        girlsVIAsums[name] += checkSum
        chexNumVIA +=1
        checkSumVIA.append(checkSum)
        if checkSum > maxValVIA:  # The maximum check sum
            maxValVIA = round(checkSum, 0)
            maxValVIADate = sh_IP_chex.cell(row=i, column=3).value
        for cd in dateChexVIA.keys():  # Counts checks by entr. and days
            checkDate = int(sh_IP_chex.cell(row=i, column=3).value[:2])
            if checkDate == int(cd[:2]):
                dateChexVIA[cd] += 1
    elif ent == entreList[1]:
        girlsRIEchex[name] += 1
        girlsRIEorders[name] += order
        girlsRIEsums[name] += checkSum
        chexNumRIE +=1
        checkSumRIE.append(checkSum)
        if checkSum > maxValRIE:  # The maximum check sum
            maxValRIE = round(checkSum, 0)
            maxValRIEDate = sh_IP_chex.cell(row=i, column=3).value
        for cd in dateChexRIE.keys():  # Counts checks by entr. and days
            checkDate = int(sh_IP_chex.cell(row=i, column=3).value[:2])
            if checkDate == int(cd[:2]):
                dateChexRIE[cd] += 1
    elif ent == entreList[2]:
        girlsSHIIchex[name] += 1
        girlsSHIIorders[name] += order
        girlsSHIIsums[name] += checkSum  # No int()
        chexNumSHII +=1
        checkSumSHII.append(checkSum)
        if checkSum > maxValSHII:  # The maximum check sum
            maxValSHII = round(checkSum, 0)
            maxValSHIIDate = sh_IP_chex.cell(row=i, column=3).value
        for cd in dateChexSHII.keys():  # Counts checks by entr. and days
            checkDate = int(sh_IP_chex.cell(row=i, column=3).value[:2])
            if checkDate == int(cd[:2]):
                dateChexSHII[cd] += 1

# Mean, median, mode
def mean(sumList):
    return sum(sumList) / len(sumList)
checkMeanVIA = int(mean(checkSumVIA))
checkMeanRIE = int(mean(checkSumRIE))
checkMeanSHII = int(mean(checkSumSHII))

checkMedianVIA = statistics.median(checkSumVIA)
checkMedianRIE = statistics.median(checkSumRIE)
checkMedianSHII = statistics.median(checkSumSHII)

checkSumVIArounded = []  # Rounding check sums for finding the mode value
for i in checkSumVIA:
    v = round(i/25)*25
    checkSumVIArounded.append(v)

checkSumRIErounded = []
for i in checkSumRIE:
    v = round(i/25)*25
    checkSumRIErounded.append(v)

checkSumSHIIrounded = []
for i in checkSumSHII:
    v = round(i/25)*25
    checkSumSHIIrounded.append(v)

checkModeVIA = statistics.mode(checkSumVIArounded)
checkModeRIE = statistics.mode(checkSumRIErounded)
checkModeSHII = statistics.mode(checkSumSHIIrounded)

print(checkMedianVIA, checkMedianRIE, checkMedianSHII)
print(checkModeVIA, checkModeRIE, checkModeSHII)

# Sorting girls...chex by values
valSort_girlsVIAchex = sorted(girlsVIAchex.values(), reverse=True)
girlsVIAchexSorted = {}
for i in valSort_girlsVIAchex:
    for k in girlsVIAchex.keys():
        if girlsVIAchex[k] == i:
            girlsVIAchexSorted[k] = girlsVIAchex[k]

valSort_girlsRIEchex = sorted(girlsRIEchex.values(), reverse=True)
girlsRIEchexSorted = {}
for i in valSort_girlsRIEchex:
    for k in girlsRIEchex.keys():
        if girlsRIEchex[k] == i:
            girlsRIEchexSorted[k] = girlsRIEchex[k]

valSort_girlsSHIIchex = sorted(girlsSHIIchex.values(), reverse=True)
girlsSHIIchexSorted = {}
for i in valSort_girlsSHIIchex:
    for k in girlsSHIIchex.keys():
        if girlsSHIIchex[k] == i:
            girlsSHIIchexSorted[k] = girlsSHIIchex[k]
            break

wbb1.close()
# ++++++++++++++++++++++++++++============================================================

# Import and process data from files 'Hours'

df = pd.read_excel('ИП_часы.xls', header = None)
df.to_excel('ИП_часы.xlsx', index = False, header = False)  # This file may not be needed
wbb3 = openpyxl.load_workbook('ИП_часы.xlsx')
hours_IP = wbb3.worksheets[0]
hours_IP.max_row-3
cas = 'Кассиры'
hours_VIA = 0
hours_RIE = 0
hours_SHII = 0
for i in range (1, hours_IP.max_row):
    dptName = str(hours_IP.cell(row=i, column=5).value)
    if cas in dptName and str(entreList[0]) in dptName:
        hours_VIA += hours_IP.cell(row=i, column=6).value
    elif cas in dptName and str(entreList[1]) in dptName:
        hours_RIE += hours_IP.cell(row=i, column=6).value
    elif cas in dptName and str(entreList[2]) in dptName:
        hours_SHII += hours_IP.cell(row=i, column=6).value
wbb3.close()
os.remove('ИП_часы.xlsx')


# wbb4 = xlsxwriter.Workbook('Нагрузка кассы.xlsx')
# wsh4 = wbb4.add_worksheet(period)
# wsh4.set_margins(left=0.6, right=0.2, top=0.2, bottom=0.2)
# wbb4.close()

workloadFile = 'Нагрузка кассы.xlsx'
try:
    wbb4 = openpyxl.load_workbook(workloadFile)
except:
    openpyxl.Workbook()
wsh4 = wbb4.create_sheet(period)
wbb4.active = wsh4

# Cell style
boldFont = Font(bold=True, size=7)
plainFont = Font(bold=False, size=9)
alignCenter = Alignment(horizontal='center', vertical='center')
alignLeft = Alignment(horizontal='left', vertical='center')
borderLines = Side(border_style='thin', color='000000')
squareBorder = Border(top=borderLines,
                      bottom=borderLines,
                      right=borderLines,
                      left=borderLines)
cellFillYellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
cellFillGreenish = PatternFill(start_color='CCFF99', end_color='CCFF99', fill_type='solid')


# Drawing the table
for i in range(3, 7):
    for j in range(1, 15):
        wsh4.cell(row=i, column=j).border = squareBorder
        wsh4.cell(row=i, column=j).alignment = alignCenter
        if i == 3:
            wsh4.cell(row=i, column=j).font = boldFont
            wsh4.cell(row=i, column=j).fill = cellFillYellow
        if j == 1 and i != 3:
            wsh4.cell(row=i, column=j).fill = cellFillGreenish

wsh4.column_dimensions['A'].width = 23
cellList = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for j in range(2, 7):
    wsh4.row_dimensions[j].height = 23
for i in cellList:
    wsh4.column_dimensions[i].width = 11
wsh4.column_dimensions['M'].width = 15
wsh4.column_dimensions['N'].width = 14

wsh4['E2'] = 'СТАТИСТИКА ПО КАССАМ'
wsh4['E2'].font = boldFont
wsh4['A3'] = period
wsh4['A4'].alignment = alignLeft
wsh4['A5'].alignment = alignLeft
wsh4['A6'].alignment = alignLeft
wsh4['A4'] = '  ИП Вербовская И.А.'
wsh4['A5'] = '  ИП Рейн И.Э.'
wsh4['A6'] = '  ИП Ширяев И.И.'
headerList = ['Кол-во сотр.', 'Кол-во чек.', 'Заказов', 'Всего часов',
              'Чек / сотр.', 'Заказ / сотр.', 'Час / сотр.', 'Чек / час', 'Средний чек', 'Медиана', 'Мода',
              'Макс. чек', 'Дата чека']
col = 2
for i in headerList:
    wsh4.cell(row=3, column=col).value = i
    col += 1

for i in range(4, 7):
    for j in range(1, 15):
        wsh4.cell(row=i, column=j).font = plainFont


# Inserting data
wsh4['B4'] = girlsNumVIA
wsh4['B5'] = girlsNumRIE
wsh4['B6'] = girlsNumSHII
wsh4['C4'] = chexNumVIA
wsh4['C5'] = chexNumRIE
wsh4['C6'] = chexNumSHII
wsh4['D4'] = orderNumVIA
wsh4['D5'] = orderNumRIE
wsh4['D6'] = orderNumSHII
wsh4['E4'] = hours_VIA
wsh4['E5'] = hours_RIE
wsh4['E6'] = hours_SHII
wsh4['F4'] = round(chexNumVIA /  girlsNumVIA, 0)
wsh4['F5'] = round(chexNumRIE /  girlsNumRIE, 0)
wsh4['F6'] = round(chexNumSHII /  girlsNumSHII, 0)
wsh4['G4'] = round(orderNumVIA /  girlsNumVIA, 1)
wsh4['G5'] = round(orderNumRIE /  girlsNumRIE, 1)
wsh4['G6'] = round(orderNumSHII /  girlsNumSHII, 1)
wsh4['H4'] = round(hours_VIA / girlsNumVIA, 1) #'=ОКРУГЛ(E4/B4;1)'
wsh4['H5'] = round(hours_RIE / girlsNumRIE, 1) # '=ОКРУГЛ(E5/B5;1)'
wsh4['H6'] = round(hours_SHII / girlsNumSHII, 1) #'=ОКРУГЛ(E6/B6;1)'
wsh4['I4'] = round(chexNumVIA / hours_VIA, 1) #'=C4 / E4'
wsh4['I5'] = round(chexNumRIE / hours_RIE, 1) #'=C5 / E5'
wsh4['I6'] = round(chexNumSHII / hours_SHII, 1) #'=C6 / E6'
wsh4['J4'] = checkMeanVIA
wsh4['J5'] = checkMeanRIE
wsh4['J6'] = checkMeanSHII
wsh4['K4'] = round(checkMedianVIA, 0)
wsh4['K5'] = round(checkMedianRIE, 0)
wsh4['K6'] = round(checkMedianSHII, 0)
wsh4['L4'] = checkModeVIA
wsh4['L5'] = checkModeRIE
wsh4['L6'] = checkModeSHII
wsh4['M4'] = maxValVIA
wsh4['M5'] = maxValRIE
wsh4['M6'] = maxValSHII
wsh4['N4'] = maxValVIADate
wsh4['N5'] = maxValRIEDate
wsh4['N6'] = maxValSHIIDate

wsh4.cell(row=4, column=3).value="{:,}".format(chexNumVIA).replace(',', ' ')
wsh4.cell(row=5, column=3).value="{:,}".format(chexNumRIE).replace(',', ' ')
wsh4.cell(row=6, column=3).value="{:,}".format(chexNumSHII).replace(',', ' ')

wsh4.cell(row=4, column=13).value = "{:,}".format(maxValVIA).replace(',', ' ')
wsh4.cell(row=5, column=13).value = "{:,}".format(maxValRIE).replace(',', ' ')
wsh4.cell(row=6, column=13).value = "{:,}".format(maxValSHII).replace(',', ' ')



# Girls' personal stats part

# Function for returning dict. key by value
def getKeyVIA(val):
    for key, value in girlsVIAchexSorted.items():
        if val == value:
            return key

row = 10
for i in girlsVIAchexSorted.values():  # Inserting list sorted by check qty
    wsh4.cell(row=row, column=3).value = i
    wsh4.cell(row=row, column=1).value = getKeyVIA(i)
    wsh4.cell(row=row, column=4).value = girlsVIAorders.get(getKeyVIA(i))
    row += 1

# Function for returning dict. key by value
def getKeyRIE(val):
    for key, value in girlsRIEchexSorted.items():
        if val == value:
            return key

row +=2
for i in girlsRIEchexSorted.values():  # Inserting list sorted by check qty
    wsh4.cell(row=row, column=3).value = i
    wsh4.cell(row=row, column=1).value = getKeyRIE(i)
    wsh4.cell(row=row, column=4).value = girlsRIEorders.get(getKeyRIE(i))
    row += 1

# Function for returning dict. key by value
def getKeySHII(val):
    for key, value in girlsSHIIchexSorted.items():
        if val == value:
            return key

row +=2
for i in girlsSHIIchexSorted.values():  # Inserting list sorted by check qty
    wsh4.cell(row=row, column=3).value = i
    wsh4.cell(row=row, column=1).value = getKeySHII(i)
    wsh4.cell(row=row, column=4).value = girlsSHIIorders.get(getKeySHII(i))
    row += 1

# Time Intervals Chart Data
row += 2
row2 = row
wsh4.cell(row=row, column=8).value = 'БД1'
wsh4.cell(row=row, column=9).value = 'БД3'
wsh4.cell(row=row, column=10).value = 'БД4'
row += 1

for i in timeIntervalsVIA.keys():
    j = timeIntervalsVIA[i]
    k = timeIntervalsRIE[i]
    l = timeIntervalsSHII[i]
    wsh4.cell(row=row, column=7).value = i
    wsh4.cell(row=row, column=8).value = j
    wsh4.cell(row=row, column=9).value = k
    wsh4.cell(row=row, column=10).value = l
    row += 1
row -= 1


# Time Intervals Bar Chart
timeIntervalsChart = BarChart()
timeIntervalsChart.title = 'Статистика по часам'
chartCats = Reference(worksheet=wsh4,
                      min_row=row2+1, max_row=row,
                      min_col=7, max_col=7)
chartData = Reference(worksheet=wsh4,
                      min_row=row2, max_row=row,
                      min_col=8, max_col=10)
timeIntervalsChart.add_data(chartData, titles_from_data=True)
timeIntervalsChart.x_axis.title = "Временные интервалы"
timeIntervalsChart.y_axis.title = "Кол-во чеков по ИП"
timeIntervalsChart.width = 25
wsh4.add_chart(timeIntervalsChart, "A62")
timeIntervalsChart.set_categories(chartCats)


# Week Days Chart Data
row += 3
row2 = row
wsh4.cell(row=row, column=8).value = 'ВИА'
wsh4.cell(row=row, column=9).value = 'РИЭ'
wsh4.cell(row=row, column=10).value = 'ШИИ'
row += 1

for i in dateChexVIA.keys():
    j = dateChexVIA[i]
    k = dateChexRIE[i]
    l = dateChexSHII[i]
    wsh4.cell(row=row, column=7).value = i
    wsh4.cell(row=row, column=8).value = j
    wsh4.cell(row=row, column=9).value = k
    wsh4.cell(row=row, column=10).value = l
    row += 1
row -= 1

# Week Days Chart
week_daysChart = BarChart()
week_daysChart.title = 'Статистика по дням недели'
chartCats = Reference(worksheet=wsh4,
                      min_row=row2+1, max_row=row,
                      min_col=7, max_col=7)
chartData = Reference(worksheet=wsh4,
                      min_row=row2, max_row=row,
                      min_col=8, max_col=10)
week_daysChart.add_data(chartData, titles_from_data=True)
week_daysChart.x_axis.title = "Дни месяца, дни недели"
week_daysChart.y_axis.title = "Кол-во чеков по ИП"
week_daysChart.width = 25
wsh4.add_chart(week_daysChart, "A89")
week_daysChart.set_categories(chartCats)



wsh4.print_area = 'A1:I97'

wbb4.save(str(workloadFile))
wbb4.close()

print(period)
